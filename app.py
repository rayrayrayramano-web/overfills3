import re
import os
import io
import subprocess
from datetime import datetime
from flask import Flask, request, send_file, render_template, jsonify
from pdf2image import convert_from_bytes
import pytesseract
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024


def find_tesseract():
    """Find tesseract binary in common locations."""
    candidates = [
        "/usr/bin/tesseract",
        "/usr/local/bin/tesseract",
        "/nix/var/nix/profiles/default/bin/tesseract",
    ]
    # Also try 'which tesseract'
    try:
        result = subprocess.run(["which", "tesseract"], capture_output=True, text=True)
        if result.returncode == 0:
            candidates.insert(0, result.stdout.strip())
    except Exception:
        pass

    for path in candidates:
        if os.path.isfile(path):
            return path
    return "tesseract"  # fallback, hope it's in PATH


pytesseract.pytesseract.tesseract_cmd = find_tesseract()


def extract_values(pdf_bytes: bytes):
    images = convert_from_bytes(pdf_bytes, first_page=1, last_page=1)
    text = pytesseract.image_to_string(images[0])

    cups_good = None
    cups_mean = None
    recipe = None

    m = re.search(r"Cups\s*\(good\)\s*[:\-]?\s*(\d+)", text)
    if m:
        cups_good = int(m.group(1))

    m = re.search(r"Cups\s*\(mean value\)\s*[:\-]?\s*([\d.,]+)", text)
    if m:
        cups_mean = float(m.group(1).replace(",", "."))

    m = re.search(r"Recip[ie]{1,2}\s*[:\-]?\s*(.+)", text, re.IGNORECASE)
    if m:
        recipe = m.group(1).strip()

    return cups_good, cups_mean, recipe, text  # return raw text for debugging


def build_excel(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Production Summary"

    headers = ["File Name", "Recipe", "Cups (Good)", "Cups Mean Value (gr)", "Total Weight (gr)"]
    header_fill = PatternFill("solid", start_color="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    row_font = Font(name="Calibri", size=11)
    alt_fill = PatternFill("solid", start_color="D6E4F0")

    data_row = 2
    for filename, recipe, cups_good, cups_mean in rows:
        fill = alt_fill if data_row % 2 == 0 else PatternFill()
        for col, val in enumerate([filename, recipe or "", cups_good, cups_mean], 1):
            cell = ws.cell(row=data_row, column=col, value=val)
            cell.font = row_font
            cell.fill = fill
            cell.alignment = Alignment(horizontal="left" if col <= 2 else "center")

        e_cell = ws.cell(row=data_row, column=5, value=f"=C{data_row}*D{data_row}")
        e_cell.font = row_font
        e_cell.fill = fill
        e_cell.alignment = Alignment(horizontal="center")
        data_row += 1

    last_data_row = data_row - 1
    total_row = data_row + 1
    total_fill = PatternFill("solid", start_color="1F4E79")

    label = ws.cell(row=total_row, column=1, value="Weighted Average Mean Value")
    label.font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    label.fill = total_fill
    label.alignment = Alignment(horizontal="left")

    for col in [2, 3, 4]:
        ws.cell(row=total_row, column=col, value="").fill = total_fill

    avg = ws.cell(row=total_row, column=5,
                  value=f"=SUM(E2:E{last_data_row})/SUM(C2:C{last_data_row})")
    avg.font = Font(bold=True, name="Calibri", size=11)
    avg.fill = PatternFill("solid", start_color="D6E4F0")
    avg.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 22

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/debug", methods=["GET"])
def debug():
    """Check what tesseract path is being used and if it works."""
    info = {
        "tesseract_cmd": pytesseract.pytesseract.tesseract_cmd,
        "tesseract_exists": os.path.isfile(pytesseract.pytesseract.tesseract_cmd),
    }
    try:
        version = pytesseract.get_tesseract_version()
        info["tesseract_version"] = str(version)
    except Exception as e:
        info["tesseract_error"] = str(e)

    try:
        result = subprocess.run(["which", "tesseract"], capture_output=True, text=True)
        info["which_tesseract"] = result.stdout.strip()
    except Exception as e:
        info["which_error"] = str(e)

    return jsonify(info)


@app.route("/process", methods=["POST"])
def process():
    files = request.files.getlist("pdfs")
    if not files:
        return jsonify({"error": "No files uploaded"}), 400

    rows = []
    skipped = []
    debug_info = []

    for f in files:
        if not f.filename.lower().endswith(".pdf"):
            continue
        try:
            pdf_bytes = f.read()
            cups_good, cups_mean, recipe, raw_text = extract_values(pdf_bytes)

            debug_info.append({
                "file": f.filename,
                "cups_good": cups_good,
                "cups_mean": cups_mean,
                "recipe": recipe,
                "ocr_preview": raw_text[:300]
            })

            if not cups_good or not cups_mean:
                skipped.append(f.filename)
                continue

            rows.append((f.filename, recipe, cups_good, cups_mean))
        except Exception as e:
            skipped.append(f.filename)
            debug_info.append({"file": f.filename, "error": str(e)})

    if not rows:
        return jsonify({
            "error": "No valid data found in uploaded PDFs",
            "debug": debug_info
        }), 400

    excel_file = build_excel(rows)
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename = f"Overills_{timestamp}.xlsx"

    return send_file(
        excel_file,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
